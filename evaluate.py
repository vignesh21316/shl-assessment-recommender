

"""
evaluate.py — SHL Assessment Recommender Evaluation
====================================================
Computes MAP@K and Recall@K against the train set ground truth.

Usage:
    python evaluate.py                          # uses Gen_AI_Dataset.xlsx (Train-Set tab)
    python evaluate.py --data path/to/file.xlsx
    python evaluate.py --k 5                    # evaluate at K=5
    python evaluate.py --api https://...        # evaluate via live API instead of local engine
"""

import argparse
import json
import sys
import time
from typing import List, Dict

import numpy as np


# ─────────────────────────────────────────────────────────────────────────────
# METRIC FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def _norm(url: str) -> str:
    return url.rstrip("/").lower().strip()


def precision_at_k(predicted: List[str], relevant: List[str], k: int) -> float:
    """Fraction of top-K predictions that are relevant."""
    if not relevant:
        return 0.0
    pred_k = [_norm(u) for u in predicted[:k]]
    rel_set = {_norm(u) for u in relevant}
    return sum(1 for u in pred_k if u in rel_set) / k


def recall_at_k(predicted: List[str], relevant: List[str], k: int) -> float:
    """Fraction of relevant items found in top-K predictions."""
    if not relevant:
        return 0.0
    pred_set = {_norm(u) for u in predicted[:k]}
    rel_set = {_norm(u) for u in relevant}
    return len(pred_set & rel_set) / len(rel_set)


def average_precision(predicted: List[str], relevant: List[str], k: int) -> float:
    """Average Precision @ K — rewards ranking relevant items higher."""
    if not relevant:
        return 0.0
    rel_set = {_norm(u) for u in relevant}
    pred_k = [_norm(u) for u in predicted[:k]]
    hits, score = 0, 0.0
    for i, url in enumerate(pred_k, 1):
        if url in rel_set:
            hits += 1
            score += hits / i
    return score / min(len(rel_set), k)


def f1_at_k(predicted: List[str], relevant: List[str], k: int) -> float:
    p = precision_at_k(predicted, relevant, k)
    r = recall_at_k(predicted, relevant, k)
    if p + r == 0:
        return 0.0
    return 2 * p * r / (p + r)


# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_ground_truth(xlsx_path: str, sheet: str = "Train-Set") -> Dict[str, List[str]]:
    """Load query → [relevant_urls] mapping from Excel sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet]
    query_map: Dict[str, List[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        query, url = str(row[0]).strip(), str(row[1]).strip() if len(row) > 1 else ""
        if url:
            query_map.setdefault(query, []).append(url)
    return query_map


# ─────────────────────────────────────────────────────────────────────────────
# RECOMMENDATION SOURCES
# ─────────────────────────────────────────────────────────────────────────────

def get_predictions_local(query: str, k: int) -> List[str]:
    """Use local SHLRecommendationEngine."""
    from rag_engine import SHLRecommendationEngine
    if not hasattr(get_predictions_local, "_engine"):
        print("  [local] Initializing engine (first call)…", flush=True)
        get_predictions_local._engine = SHLRecommendationEngine()
        get_predictions_local._engine.initialize()
    engine = get_predictions_local._engine
    results = engine.recommend(query, max_results=k)
    return [r["url"] for r in results]


def get_predictions_api(query: str, k: int, api_url: str) -> List[str]:
    """Use deployed FastAPI endpoint."""
    import requests as req
    try:
        resp = req.post(f"{api_url}/recommend", json={"query": query}, timeout=60)
        resp.raise_for_status()
        data = resp.json()
        return [r["url"] for r in data.get("recommended_assessments", [])][:k]
    except Exception as e:
        print(f"  [api] Error: {e}", flush=True)
        return []


# ─────────────────────────────────────────────────────────────────────────────
# MAIN EVALUATION LOOP
# ─────────────────────────────────────────────────────────────────────────────

def evaluate(
    data_path: str = "Gen_AI_Dataset.xlsx",
    k: int = 10,
    api_url: str = None,
    verbose: bool = True,
    delay: float = 0.5,
) -> Dict:
    print(f"\n{'='*60}")
    print(f"  SHL Assessment Recommender — Evaluation")
    print(f"  Dataset : {data_path}")
    print(f"  K       : {k}")
    print(f"  Mode    : {'API → ' + api_url if api_url else 'Local engine'}")
    print(f"{'='*60}\n")

    ground_truth = load_ground_truth(data_path)
    print(f"Loaded {len(ground_truth)} queries from train set.\n")

    records = []
    for i, (query, relevant) in enumerate(ground_truth.items(), 1):
        print(f"[{i:02d}/{len(ground_truth)}] {query[:80]}…")

        if api_url:
            predicted = get_predictions_api(query, k, api_url)
            if delay and i < len(ground_truth):
                time.sleep(delay)
        else:
            predicted = get_predictions_local(query, k)

        ap  = average_precision(predicted, relevant, k)
        rec = recall_at_k(predicted, relevant, k)
        pre = precision_at_k(predicted, relevant, k)
        f1  = f1_at_k(predicted, relevant, k)

        records.append({"query": query, "ap": ap, "recall": rec, "precision": pre, "f1": f1,
                         "n_relevant": len(relevant), "n_predicted": len(predicted)})

        if verbose:
            print(f"        AP@{k}={ap:.3f}  Recall@{k}={rec:.3f}  P@{k}={pre:.3f}  F1={f1:.3f}"
                  f"  (rel={len(relevant)}, pred={len(predicted)})")

    # Aggregate
    map_k    = float(np.mean([r["ap"]        for r in records]))
    mrec_k   = float(np.mean([r["recall"]    for r in records]))
    mpre_k   = float(np.mean([r["precision"] for r in records]))
    mf1_k    = float(np.mean([r["f1"]        for r in records]))

    print(f"\n{'='*60}")
    print(f"  RESULTS @ K={k}")
    print(f"{'='*60}")
    print(f"  MAP@{k:<4}     {map_k:.4f}")
    print(f"  Recall@{k:<3}   {mrec_k:.4f}")
    print(f"  Precision@{k}  {mpre_k:.4f}")
    print(f"  F1@{k:<4}      {mf1_k:.4f}")
    print(f"  Queries       {len(records)}")
    print(f"{'='*60}\n")

    summary = {
        "k": k, "n_queries": len(records),
        f"MAP@{k}": round(map_k, 4),
        f"Recall@{k}": round(mrec_k, 4),
        f"Precision@{k}": round(mpre_k, 4),
        f"F1@{k}": round(mf1_k, 4),
        "per_query": records,
    }

    # Save results
    out = f"eval_results_k{k}.json"
    with open(out, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"Full results saved to {out}")

    return summary


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Evaluate SHL Assessment Recommender")
    parser.add_argument("--data", default="Gen_AI_Dataset.xlsx",
                        help="Path to dataset Excel file (default: Gen_AI_Dataset.xlsx)")
    parser.add_argument("--k", type=int, default=10,
                        help="K for MAP@K and Recall@K (default: 10)")
    parser.add_argument("--api", default=None,
                        help="API base URL to evaluate against (e.g. https://shl-recommender-manikanta.onrender.com). "
                             "Omit to use local engine.")
    parser.add_argument("--quiet", action="store_true",
                        help="Suppress per-query output")
    parser.add_argument("--delay", type=float, default=0.5,
                        help="Seconds between API calls (default: 0.5)")
    args = parser.parse_args()

    results = evaluate(
        data_path=args.data,
        k=args.k,
        api_url=args.api,
        verbose=not args.quiet,
        delay=args.delay,
    )
    sys.exit(0)
